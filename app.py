import streamlit as st
import pandas as pd
import json
import os
import re
import tempfile
from io import BytesIO

from pdf_statement_reader.parse import parse_statement  # noqa: direct import to avoid __init__.py loading pikepdf

BANK_CONFIGS = {
    "ABSA": "za.absa.transaction_history",
    "FNB": "za.fnb.business",
    "Standard Bank": "za.standardbank.current",
    "Standard Bank (Tax Invoice)": "za.standardbank.tax_invoice",
    "Nedbank": "za.nedbank.cheque",
}

PASTEL_MAX_CHARS = 36

UNREADABLE_PLACEHOLDER = "[DESCRIPTION UNREADABLE IN PDF - CHECK MANUALLY]"

# Matches a bare card/reference pattern like "428104*6314 12 Mar" — indicates
# the actual description was not extractable from the PDF text layer.
_CARD_REF_ONLY_RE = re.compile(
    r"^\d+\*\d+\s+\d{1,2}\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*$",
    re.IGNORECASE,
)

# Transaction type prefixes to strip (longest first so longer matches win)
STRIP_PREFIXES = [
    # Standard Bank
    "REAL TIME TRANSFER FROM ",
    "REAL TIME TRANSFER TO ",
    "CREDIT TRANSFER ",
    "IB PAYMENT FROM ",
    "IB PAYMENT TO ",
    "LOAN REPAYMENT ",
    # ABSA
    "ACB CREDIT MERCH/SERV ",
    "ACB DEBIT:INTERNAL MERCH/SERV ",
    "ACB DEBIT:EXTERNAL ",
    "DIGITAL PAYMENT DT ",
    "IMDTE DIGITAL PMT ",
    "DIGITAL TRANSF CR ",
    "IMMEDIATE TRF CR ",
    "CARDLESS CASH DEP ",
    "NOTIFIC FEE SMS ",
    "PROOF OF PMT EMAIL ",
    "PROOF OF PAYMT SMS ",
    "ARCHIVE STMT ENQ ",
    "STAMPED STATEMENT ",
    "ACB CREDIT ",
    "ACB DEBIT ",
    # Standard Bank Tax Invoice
    "MAGTAPE CREDIT ",
    "MAGTAPE DEBIT ",
    "BOLSA THIRD PARTY PAYMENT ",
    "Inward Swift ",
    # FNB
    "Magtape Debit ",
    "Magtape Credit ",
    "POS Purchase ",
    "Rtc Credit ",
    "FNB OB Pmt ",
    # Nedbank
    "Instant payment fee",
]


def clean_description_for_pastel(desc):
    """Clean a bank description to fit Pastel's 36-character ledger field.

    Strips transaction type prefixes, fee amounts in parentheses,
    account/card/reference numbers, and trailing dates to keep only
    the meaningful recipient or expense name.
    """
    if not desc or desc in ("Bank Fee/Charge", "nan", "", UNREADABLE_PLACEHOLDER):
        return desc

    # Strip known transaction type prefixes
    for prefix in STRIP_PREFIXES:
        if desc.upper().startswith(prefix.upper()):
            remainder = desc[len(prefix):].strip()
            if remainder:  # only strip if something meaningful remains
                desc = remainder
            break

    # FNB: strip "FNB OB Pmt FNB OB 000000942 Sar " pattern (prefix + ref + short code)
    desc = re.sub(r"^FNB OB Pmt FNB OB \d+ \w{3}\s+", "", desc, flags=re.IGNORECASE)

    # Remove fee amounts in parentheses like "( 10,00 )" or "(19,75 )"
    desc = re.sub(r"\(\s*[\d,\.]+\s*\)\s*", "", desc)

    # Remove "ABSA BANK" after stripping ABSA payment prefixes
    desc = re.sub(r"^ABSA BANK\s+", "", desc)
    # Also handle "ABSA BANK" preceded by space (from DIGITAL PAYMENT DT cleanup)
    desc = re.sub(r"\s*ABSA BANK\s*", " ", desc)

    # Remove EFFEC date references like "(EFFEC 30012026)"
    desc = re.sub(r"\(EFFEC\s*\d*\)", "", desc)

    # Remove card numbers glued to text (e.g. "GERMI5412820023910484" → "GERMI")
    desc = re.sub(r"(\D)\d{10,}", r"\1", desc)

    # Remove standalone long digit sequences (card/account numbers, 8+ digits)
    desc = re.sub(r"\b\d{8,}\b", "", desc)

    # Remove hex reference codes like "161A2AEED7"
    desc = re.sub(r"\b[0-9A-F]{10,}\b", "", desc)

    # Remove Nedbank ADT-style reference codes glued to text (e.g. "2114879792ADTA059137")
    desc = re.sub(r"\d{5,}[A-Z]+\d+", "", desc)

    # Remove POS card references like "491050*0796"
    desc = re.sub(r"\d+\*\d+", "", desc)

    # Remove card suffixes DD/CC at end
    desc = re.sub(r"\s+(DD|CC)\s*$", "", desc)

    # Remove trailing short dates like "02 Jan"
    desc = re.sub(
        r"\s+\d{1,2}\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*$",
        "",
        desc,
        flags=re.IGNORECASE,
    )

    # Remove "ACC" followed by account number
    desc = re.sub(r"\bACC\s+\d+\b", "", desc)

    # Remove standalone short reference numbers (3-7 digits) at end of line
    desc = re.sub(r"\s+\d{3,7}\s*$", "", desc)

    # Remove "NOTIFICATION" if duplicated (e.g. "... 0013 NOTIFICATION")
    desc = re.sub(r"\s+NOTIFICATION\s*$", "", desc, flags=re.IGNORECASE)

    # Remove percentage rates like "@17750%" or "@17,750%"
    desc = re.sub(r"@[\d,]+%", "", desc)

    # Clean up multiple spaces and trim
    desc = re.sub(r"\s+", " ", desc).strip()

    # Remove trailing/leading dashes and hyphens
    desc = desc.strip(" -")

    return desc


def build_pastel_csv(df):
    """Convert bank-specific DataFrame to Pastel format: Date, Description, Amount."""
    pastel_df = pd.DataFrame()

    # Date
    date_col = [c for c in df.columns if "date" in c.lower()][0]
    pastel_df["Date"] = df[date_col].dt.strftime("%Y/%m/%d")

    # Description (full original)
    desc_col = [
        c
        for c in df.columns
        if any(k in c.lower() for k in ["description", "reference", "transaction"])
    ][0]
    raw_desc = df[desc_col].fillna("").astype(str).replace("nan", "")

    # Amount — merge In/Out or Debit/Credit if split into separate columns
    if "Amount" in df.columns:
        pastel_df["Amount"] = df["Amount"]
    elif "In (R)" in df.columns and "Out (R)" in df.columns:
        pastel_df["Amount"] = df["In (R)"].fillna(0) + df["Out (R)"].fillna(0)
    elif "Debits" in df.columns and "Credits" in df.columns:
        pastel_df["Amount"] = df["Credits"].fillna(0) - df["Debits"].fillna(0).abs()
    else:
        amt_col = [
            c
            for c in df.columns
            if "amount" in c.lower() or "debit" in c.lower()
        ][0]
        pastel_df["Amount"] = df[amt_col]

    # Resolve missing/unreadable descriptions:
    # - Empty + small negative amount (|amt| < 200) → "Bank Fee/Charge" (typical FNB fee)
    # - Empty/card-ref-only + anything else → UNREADABLE_PLACEHOLDER (PDF rendering bug)
    def resolve_description(desc, amount):
        desc = str(desc).strip()
        is_card_ref_only = bool(_CARD_REF_ONLY_RE.match(desc))
        if not desc or is_card_ref_only:
            try:
                amt = float(amount)
            except (ValueError, TypeError):
                amt = 0
            if not desc and amt < 0 and abs(amt) < 200:
                return "Bank Fee/Charge"
            return UNREADABLE_PLACEHOLDER
        return desc

    pastel_df["Description"] = [
        resolve_description(d, a) for d, a in zip(raw_desc, pastel_df["Amount"])
    ]

    # Ledger Description (cleaned, max 36 chars for Pastel)
    pastel_df["Ledger Description"] = pastel_df["Description"].apply(
        clean_description_for_pastel
    )

    # Filter out non-transaction rows
    skip_patterns = [
        "BALANCE BROUGHT FORWARD", "OPENING BALANCE", "CLOSING BALANCE",
        "CARRIED FORWARD", "BROUGHT FORWARD", "PROVISIONAL STATEMENT",
    ]
    mask = pastel_df["Description"].apply(
        lambda d: not any(p in str(d).upper() for p in skip_patterns)
    )
    pastel_df = pastel_df[mask].reset_index(drop=True)

    # Drop rows with no valid date
    pastel_df = pastel_df[pastel_df["Date"].notna() & (pastel_df["Date"] != "NaT")].reset_index(drop=True)

    return pastel_df


def highlight_row(row):
    """Highlight rows: red for unreadable descriptions, yellow for over-limit ledger descriptions."""
    if str(row["Description"]) == UNREADABLE_PLACEHOLDER:
        return ["background-color: #FFB3B3"] * len(row)
    if len(str(row["Ledger Description"])) > PASTEL_MAX_CHARS:
        return ["background-color: #FFFF00"] * len(row)
    return [""] * len(row)


def to_excel_with_highlights(pastel_df):
    """Create an Excel file with red highlighting for unreadable descriptions and yellow for over-limit rows."""
    from openpyxl.styles import PatternFill

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pastel_df.to_excel(writer, index=False, sheet_name="Transactions")
        ws = writer.sheets["Transactions"]
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")

        desc_col_idx = list(pastel_df.columns).index("Description") + 1
        ledger_col_idx = list(pastel_df.columns).index("Ledger Description") + 1

        for row_idx in range(2, len(pastel_df) + 2):
            desc_cell = ws.cell(row=row_idx, column=desc_col_idx)
            ledger_cell = ws.cell(row=row_idx, column=ledger_col_idx)
            if str(desc_cell.value) == UNREADABLE_PLACEHOLDER:
                for col_idx in range(1, len(pastel_df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = red_fill
            elif ledger_cell.value and len(str(ledger_cell.value)) > PASTEL_MAX_CHARS:
                for col_idx in range(1, len(pastel_df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    return output.getvalue()


def load_config(config_spec):
    local_dir = os.path.join(os.path.dirname(__file__), "pdf_statement_reader")
    config_dir = os.path.join(*config_spec.split(".")[:-1])
    config_file = config_spec.split(".")[-1] + ".json"
    config_path = os.path.join(local_dir, "config", config_dir, config_file)
    with open(config_path) as f:
        return json.load(f)


# --- Page config ---
st.set_page_config(
    page_title="Bank Statement to CSV", page_icon="🏦", layout="centered"
)

st.title("Bank Statement to CSV")
st.caption(
    "Convert PDF bank statements into Sage Pastel-compatible CSV files. "
    "Descriptions are automatically cleaned to fit Pastel's 36-character ledger limit."
)

# --- Inputs ---
col1, col2 = st.columns([1, 2])
with col1:
    bank = st.selectbox("Bank", list(BANK_CONFIGS.keys()))
with col2:
    uploaded_file = st.file_uploader(
        "Upload PDF statement", type=["pdf"], label_visibility="visible"
    )

if not uploaded_file:
    st.info("Upload a PDF bank statement to get started.")
    st.stop()

# --- Process ---
with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
    tmp.write(uploaded_file.read())
    tmp_path = tmp.name

try:
    with st.spinner("Extracting transactions..."):
        config = load_config(BANK_CONFIGS[bank])
        df = parse_statement(tmp_path, config)
        pastel_df = build_pastel_csv(df)

    st.success(f"Extracted **{len(pastel_df)}** transactions from **{bank}** statement.")

    # Count unreadable descriptions (PDF rendering issue — text stored as graphics)
    n_unreadable = (pastel_df["Description"] == UNREADABLE_PLACEHOLDER).sum()
    if n_unreadable > 0:
        st.error(
            f"**{n_unreadable}** row(s) have descriptions that couldn't be extracted from the PDF "
            "(highlighted red). These are rendered as graphics in the source PDF — open the PDF "
            "and type the descriptions in manually. Amounts and dates are correct."
        )

    # Count descriptions over the limit
    over_limit = pastel_df["Ledger Description"].apply(lambda x: len(str(x)) > PASTEL_MAX_CHARS)
    n_over = over_limit.sum()
    if n_over > 0:
        st.warning(
            f"**{n_over}** description(s) still exceed {PASTEL_MAX_CHARS} characters "
            f"(highlighted in yellow). Review these before importing into Pastel."
        )

    # --- Preview with highlighting ---
    st.subheader("Preview")
    styled = pastel_df.style.apply(highlight_row, axis=1)
    st.dataframe(styled, use_container_width=True, hide_index=True)

    # --- Downloads ---
    st.subheader("Download")
    col_csv, col_xlsx = st.columns(2)
    file_stem = uploaded_file.name.replace(".pdf", "").replace(".PDF", "")

    with col_csv:
        csv_data = pastel_df.to_csv(index=False, float_format="%.2f")
        st.download_button(
            label="Download CSV",
            data=csv_data,
            file_name=f"{file_stem}.csv",
            mime="text/csv",
            type="primary",
        )

    with col_xlsx:
        xlsx_data = to_excel_with_highlights(pastel_df)
        st.download_button(
            label="Download Excel (highlighted)",
            data=xlsx_data,
            file_name=f"{file_stem}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # --- Raw data toggle ---
    with st.expander("Show raw extracted data"):
        st.dataframe(df, use_container_width=True, hide_index=True)

except Exception as e:
    st.error(f"Failed to parse statement: {e}")
    st.caption(
        "Make sure you selected the correct bank and uploaded a valid PDF statement."
    )

finally:
    os.unlink(tmp_path)
